import hashlib
import secrets
import io
from datetime import datetime
from contextlib import asynccontextmanager
from fastapi import FastAPI, Depends, HTTPException
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from sqlalchemy.orm import Session
from sqlalchemy import func
from database import get_db, engine
import models
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

models.Base.metadata.create_all(bind=engine)

security = HTTPBearer()


def seed_initial_data(db: Session):
    if db.query(models.User).count() > 0:
        return
    users = [
        ("admin", "admin123", "Administrador", "admin"),
        ("juan", "almacen123", "Juan Pérez", "admin"),
    ]
    for username, password, name, role in users:
        h = hashlib.sha256(password.encode()).hexdigest()
        db.add(models.User(username=username, password_hash=h, name=name, role=role))
    db.commit()


@asynccontextmanager
async def lifespan(app: FastAPI):
    db = next(get_db())
    try:
        seed_initial_data(db)
    finally:
        db.close()
    yield


app = FastAPI(title="Control de Acero - Inventario", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def get_current_user(
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db),
):
    token_str = credentials.credentials
    token = db.query(models.Token).filter(models.Token.token == token_str).first()
    if not token:
        raise HTTPException(status_code=401, detail="Token inválido")
    user = db.query(models.User).filter(models.User.id == token.user_id).first()
    if not user:
        raise HTTPException(status_code=401, detail="Usuario no encontrado")
    return user


# ---------- AUTH ----------

@app.post("/api/login")
def login(data: dict, db: Session = Depends(get_db)):
    username = data.get("username", "")
    password = data.get("password", "")
    h = hashlib.sha256(password.encode()).hexdigest()
    user = db.query(models.User).filter(models.User.username == username).first()
    if not user or user.password_hash != h:
        raise HTTPException(status_code=400, detail="Credenciales inválidas")
    token_str = secrets.token_hex(32)
    token = models.Token(token=token_str, user_id=user.id)
    db.add(token)
    db.commit()
    return {
        "token": token_str,
        "user": {"id": user.id, "name": user.name, "username": user.username, "role": user.role},
    }


@app.post("/api/logout")
def logout(
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db),
):
    db.query(models.Token).filter(models.Token.token == credentials.credentials).delete()
    db.commit()
    return {"ok": True}


@app.get("/api/me")
def get_me(user=Depends(get_current_user)):
    return {"id": user.id, "name": user.name, "username": user.username, "role": user.role}


# ---------- INVENTORY ----------

@app.get("/api/inventory")
def get_inventory(user=Depends(get_current_user), db: Session = Depends(get_db)):
    entries = (
        db.query(
            models.Movement.steel_type_id,
            func.sum(models.Movement.quantity).label("total"),
        )
        .filter(models.Movement.movement_type == "entry")
        .group_by(models.Movement.steel_type_id)
        .subquery()
    )
    exits = (
        db.query(
            models.Movement.steel_type_id,
            func.sum(models.Movement.quantity).label("total"),
        )
        .filter(models.Movement.movement_type == "exit")
        .group_by(models.Movement.steel_type_id)
        .subquery()
    )

    steel_types = db.query(models.SteelType).order_by(models.SteelType.name).all()
    result = []
    for st in steel_types:
        entry_qty = db.query(entries.c.total).filter(entries.c.steel_type_id == st.id).scalar() or 0
        exit_qty = db.query(exits.c.total).filter(exits.c.steel_type_id == st.id).scalar() or 0
        stock = entry_qty - exit_qty

        last_movement = (
            db.query(models.Movement)
            .filter(models.Movement.steel_type_id == st.id)
            .order_by(models.Movement.created_at.desc())
            .first()
        )
        last_info = None
        if last_movement:
            u = db.query(models.User).filter(models.User.id == last_movement.registered_by).first()
            last_info = {
                "date": last_movement.created_at.isoformat() if last_movement.created_at else "",
                "by": u.name if u else "Desconocido",
                "type": last_movement.movement_type,
            }

        result.append({
            "id": st.id,
            "name": st.name,
            "stock": stock,
            "last_movement": last_info,
        })
    return result


# ---------- MOVEMENTS ----------

@app.get("/api/movements")
def list_movements(user=Depends(get_current_user), db: Session = Depends(get_db)):
    q = (
        db.query(models.Movement)
        .order_by(models.Movement.created_at.desc())
        .limit(200)
        .all()
    )
    result = []
    for m in q:
        u = db.query(models.User).filter(models.User.id == m.registered_by).first()
        result.append({
            "id": m.id,
            "steel_type_id": m.steel_type_id,
            "steel_type_name": m.steel_type.name if m.steel_type else "Desconocido",
            "movement_type": m.movement_type,
            "quantity": m.quantity,
            "person_name": m.person_name or "",
            "note": m.note or "",
            "registered_by_name": u.name if u else "Desconocido",
            "created_at": m.created_at.isoformat() if m.created_at else "",
        })
    return result


@app.post("/api/movements/entry")
def create_entry(
    data: dict,
    user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    steel_type_id = data.get("steel_type_id")
    quantity = data.get("quantity", 0)
    note = data.get("note", "")
    if not steel_type_id or quantity <= 0:
        raise HTTPException(status_code=400, detail="Datos inválidos")
    st = db.query(models.SteelType).filter(models.SteelType.id == steel_type_id).first()
    if not st:
        raise HTTPException(status_code=404, detail="Tipo de acero no encontrado")
    m = models.Movement(
        steel_type_id=steel_type_id,
        movement_type="entry",
        quantity=quantity,
        note=note,
        registered_by=user.id,
    )
    db.add(m)
    db.commit()
    db.refresh(m)
    return {"ok": True, "id": m.id}


@app.post("/api/movements/exit")
def create_exit(
    data: dict,
    user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    steel_type_id = data.get("steel_type_id")
    quantity = data.get("quantity", 0)
    person_name = data.get("person_name", "").strip()
    note = data.get("note", "")
    if not steel_type_id or quantity <= 0:
        raise HTTPException(status_code=400, detail="Datos inválidos")
    if not person_name:
        raise HTTPException(status_code=400, detail="Nombre de quién sacó es obligatorio")
    st = db.query(models.SteelType).filter(models.SteelType.id == steel_type_id).first()
    if not st:
        raise HTTPException(status_code=404, detail="Tipo de acero no encontrado")

    entry_total = (
        db.query(func.sum(models.Movement.quantity))
        .filter(
            models.Movement.steel_type_id == steel_type_id,
            models.Movement.movement_type == "entry",
        )
        .scalar() or 0
    )
    exit_total = (
        db.query(func.sum(models.Movement.quantity))
        .filter(
            models.Movement.steel_type_id == steel_type_id,
            models.Movement.movement_type == "exit",
        )
        .scalar() or 0
    )
    stock = entry_total - exit_total
    if quantity > stock:
        raise HTTPException(status_code=400, detail=f"Stock insuficiente. Disponible: {stock}")

    m = models.Movement(
        steel_type_id=steel_type_id,
        movement_type="exit",
        quantity=quantity,
        person_name=person_name,
        note=note,
        registered_by=user.id,
    )
    db.add(m)
    db.commit()
    db.refresh(m)
    return {"ok": True, "id": m.id}


@app.delete("/api/movements/{movement_id}")
def delete_movement(
    movement_id: int,
    user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores")
    m = db.query(models.Movement).filter(models.Movement.id == movement_id).first()
    if not m:
        raise HTTPException(status_code=404, detail="Movimiento no encontrado")
    db.delete(m)
    db.commit()
    return {"ok": True}


@app.get("/api/inventory/export")
def export_inventory(user=Depends(get_current_user), db: Session = Depends(get_db)):
    entries = (
        db.query(
            models.Movement.steel_type_id,
            func.sum(models.Movement.quantity).label("total"),
        )
        .filter(models.Movement.movement_type == "entry")
        .group_by(models.Movement.steel_type_id)
        .subquery()
    )
    exits = (
        db.query(
            models.Movement.steel_type_id,
            func.sum(models.Movement.quantity).label("total"),
        )
        .filter(models.Movement.movement_type == "exit")
        .group_by(models.Movement.steel_type_id)
        .subquery()
    )

    steel_types = db.query(models.SteelType).order_by(models.SteelType.name).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4f46e5", end_color="4f46e5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["Tipo de Acero", "Cantidad Disponible", "Estado"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 15

    row = 2
    for st in steel_types:
        entry_qty = db.query(entries.c.total).filter(entries.c.steel_type_id == st.id).scalar() or 0
        exit_qty = db.query(exits.c.total).filter(exits.c.steel_type_id == st.id).scalar() or 0
        stock = entry_qty - exit_qty
        status = "Bajo stock" if stock < 5 else "OK"

        ws.cell(row=row, column=1, value=st.name).border = thin_border
        ws.cell(row=row, column=2, value=stock).border = thin_border
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        status_cell = ws.cell(row=row, column=3, value=status)
        status_cell.border = thin_border
        status_cell.alignment = Alignment(horizontal="center")
        if stock < 5:
            status_cell.font = Font(color="EF4444", bold=True)
        row += 1

    ws2 = wb.create_sheet("Movimientos")
    mov_headers = [
        "Fecha", "Tipo", "Tipo de Acero", "Cantidad",
        "Quién Registró", "Quién Sacó", "Nota",
    ]
    for col, h in enumerate(mov_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    widths = [20, 12, 30, 12, 20, 20, 30]
    for i, w in enumerate(widths, 1):
        ws2.column_dimensions[chr(64 + i)].width = w

    movements = (
        db.query(models.Movement)
        .order_by(models.Movement.created_at.desc())
        .limit(500)
        .all()
    )
    row = 2
    for m in movements:
        u = db.query(models.User).filter(models.User.id == m.registered_by).first()
        ws2.cell(row=row, column=1, value=m.created_at.strftime("%d/%m/%Y %H:%M") if m.created_at else "").border = thin_border
        ws2.cell(row=row, column=2, value="Entrada" if m.movement_type == "entry" else "Salida").border = thin_border
        ws2.cell(row=row, column=2).alignment = Alignment(horizontal="center")
        ws2.cell(row=row, column=3, value=m.steel_type.name if m.steel_type else "").border = thin_border
        ws2.cell(row=row, column=4, value=m.quantity).border = thin_border
        ws2.cell(row=row, column=4).alignment = Alignment(horizontal="center")
        ws2.cell(row=row, column=5, value=u.name if u else "").border = thin_border
        ws2.cell(row=row, column=6, value=m.person_name or "").border = thin_border
        ws2.cell(row=row, column=7, value=m.note or "").border = thin_border
        row += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inventario_acero.xlsx"},
    )


# ---------- STEEL TYPES ----------

@app.get("/api/steel-types")
def list_steel_types(user=Depends(get_current_user), db: Session = Depends(get_db)):
    q = db.query(models.SteelType).order_by(models.SteelType.name).all()
    return [{"id": st.id, "name": st.name} for st in q]


@app.post("/api/steel-types")
def create_steel_type(
    data: dict,
    user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores")
    name = data.get("name", "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Nombre requerido")
    existing = db.query(models.SteelType).filter(models.SteelType.name == name).first()
    if existing:
        raise HTTPException(status_code=400, detail="Ya existe")
    st = models.SteelType(name=name)
    db.add(st)
    db.commit()
    db.refresh(st)
    return {"id": st.id, "name": st.name}


@app.delete("/api/steel-types/{steel_type_id}")
def delete_steel_type(
    steel_type_id: int,
    user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores")
    st = db.query(models.SteelType).filter(models.SteelType.id == steel_type_id).first()
    if not st:
        raise HTTPException(status_code=404)
    count = db.query(models.Movement).filter(models.Movement.steel_type_id == steel_type_id).count()
    if count > 0:
        raise HTTPException(status_code=400, detail="No se puede eliminar un tipo con movimientos")
    db.delete(st)
    db.commit()
    return {"ok": True}


# ---------- USERS ----------

@app.get("/api/users")
def list_users(user=Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores")
    q = db.query(models.User).all()
    return [{"id": u.id, "name": u.name, "username": u.username, "role": u.role} for u in q]


@app.post("/api/users")
def create_user(data: dict, user=Depends(get_current_user), db: Session = Depends(get_db)):
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores")
    username = data.get("username", "").strip()
    password = data.get("password", "")
    name = data.get("name", "").strip()
    role = data.get("role", "admin")
    if not username or not password or not name:
        raise HTTPException(status_code=400, detail="Todos los campos son obligatorios")
    existing = db.query(models.User).filter(models.User.username == username).first()
    if existing:
        raise HTTPException(status_code=400, detail="El usuario ya existe")
    h = hashlib.sha256(password.encode()).hexdigest()
    u = models.User(username=username, password_hash=h, name=name, role=role)
    db.add(u)
    db.commit()
    db.refresh(u)
    return {"id": u.id, "name": u.name, "username": u.username, "role": u.role}


@app.delete("/api/users/{user_id}")
def delete_user(
    user_id: int,
    user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores")
    if user.id == user_id:
        raise HTTPException(status_code=400, detail="No puedes eliminarte a ti mismo")
    u = db.query(models.User).filter(models.User.id == user_id).first()
    if not u:
        raise HTTPException(status_code=404)
    db.delete(u)
    db.commit()
    return {"ok": True}


@app.put("/api/users/{user_id}/password")
def change_password(
    user_id: int,
    data: dict,
    user=Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores")
    u = db.query(models.User).filter(models.User.id == user_id).first()
    if not u:
        raise HTTPException(status_code=404)
    new_password = data.get("password", "")
    if not new_password or len(new_password) < 4:
        raise HTTPException(status_code=400, detail="Contraseña debe tener al menos 4 caracteres")
    u.password_hash = hashlib.sha256(new_password.encode()).hexdigest()
    db.commit()
    return {"ok": True}


# ---------- FRONTEND ----------

app.mount("/static", StaticFiles(directory="frontend"), name="static")


@app.get("/")
def index():
    return FileResponse("frontend/index.html")


@app.exception_handler(404)
async def not_found(request, exc):
    if request.url.path.startswith("/api/"):
        return JSONResponse({"detail": "Not found"}, status_code=404)
    return FileResponse("frontend/index.html")
